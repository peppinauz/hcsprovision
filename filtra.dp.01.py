## LINUX#!/usr/bin/python3
## OSX
#!/usr/local/bin/python3


import csv
import openpyxl
import time
import sys
import json



if len(sys.argv) != 5:
    print("ERROR: python3 <nombre-fichero.py> <cmo site data> <ClusterXX> XXXX <logconfigfile>")
    print(">>>>  [1] <cmo site data>")
    print(">>>>  [2] cluster path:  ../CMO/ClusterXX")
    print(">>>>  [3] XXXX: SLC")
    print(">>>>  [4] config log file: ../FMO/XXXX/logconfigfilename-.txt")
    exit(1)
else:
    outconfigfile=sys.argv[1]   ## CMO configuration
    cluster=sys.argv[2]         ## ClusterXX
    slc=sys.argv[3]             ## INPUTS: SLC
    logconfigfile=sys.argv[4]   ## LOG config file

data = {}
data['dp'] = []
data['loc']=[]
data['mrgl'] = []
data['mrg'] = []
data['srst'] = []
data['gw'] = []
data['cnf'] = []
data['mtp'] = []
data['trans'] = []
data['e164'] = []
data['agencia'] = []
data['devices'] = []
data['fmosite'] = []
data['huntpilot']=[]
data['cpgnumber']=[]
data['sipptest']=[]

## Variables
sitename=""
ramais=""
epnm=""
areacode=""
ac=""
cc=""
intrasite1=""
intrasite2=""
patternintra=""

def busca_digito(fila,intra1,intra2,max):
    #print("**",row['Directory Number 1'],row['Directory Number 2'],row['Directory Number 3'],row['Directory Number 4'],row['Directory Number 5'],row['Directory Number 6'])
    #
    # PRIMER dígito
    #
    if row['Directory Number 1'][5:6] not in intra1:
        intra1=intra1+row['Directory Number 1'][5:6]
        #print(">> ",intra1)

    if max >= 2:
        if row['Directory Number 2'][5:6] not in intra1:
            intra1=intra1+row['Directory Number 2'][5:6]
            #print(">> ",intra1)

    if max >= 3:
        if row['Directory Number 3'][5:6] not in intra1:
            intra1=intra1+row['Directory Number 3'][5:6]
            #print(">> ",intra1)

    if max >= 4:
        if row['Directory Number 4'][5:6] not in intra1:
            intra1=intra1+row['Directory Number 3'][5:6]
            #print(">> ",intra1)

    if max >= 5:
        if row['Directory Number 5'][5:6] not in intra1:
            intra1=intra1+row['Directory Number 3'][5:6]
            #print(">> ",intra1)

    if max >= 6:
        if row['Directory Number 6'][5:6] not in intra1:
            intra1=intra1+row['Directory Number 3'][5:6]
            #print(">> ",intra1)

    #
    # SEGUNDO dígito
    #
    if row['Directory Number 1'][6:7] not in intra2 :
        intra2=intra2+row['Directory Number 1'][6:7]
        #print(">> ",intra1)

    if max >= 2:
        if row['Directory Number 2'][6:7] not in intra2:
            intra2=intra2+row['Directory Number 2'][6:7]
            #print(">> ",intra1)

    if max >= 3:
        if row['Directory Number 3'][6:7] not in intra2:
            intra2=intra2+row['Directory Number 3'][6:7]
            #print(">> ",intra1)

    if max >= 4:
        if row['Directory Number 4'][6:7] not in intra2:
            intra2=intra2+row['Directory Number 3'][6:7]
            #print(">> ",intra1)

    if max >= 5:
        if row['Directory Number 5'][6:7] not in intra2:
            intra2=intra2+row['Directory Number 3'][6:7]
            #print(">> ",intra1)

    if max >= 6:
        if row['Directory Number 6'][6:7] not in intra2:
            intra2=intra2+row['Directory Number 3'][6:7]
            #print(">> ",intra1)

    return intra1,intra2



## Ficheros de salida
#outconfigfile="../FMO/"+slc+"/01.cmo-site-config.json"

## LOG File
f = open(logconfigfile, 'w')

print("(II) #################################################", file=f)
print("(II) #################################################", file=f)
print("(II) INPUT CMO configuration      ::",outconfigfile, file=f)
print("(II) INPUT Cluster Path           ::",cluster, file=f)
print("(II) INPUT SLC                    :: ",slc, file=f)
print("(II) INPUT LOG configuration file :: ",logconfigfile, file=f)
print("(II) Extrayendo info de CMO", file=f)

## Ficheros de entrada
dpfile=cluster+"/devicepool.csv"
dproffile=cluster+"/deviceprofile.csv"
phonefile=cluster+"/phone.csv"
srstfile=cluster+"/srst.csv"
locationfile=cluster+"/location.csv"
gwfile=cluster+"/gateway.csv"
mrglfile=cluster+"/mediaresourcegrouplist.csv"
mrgfile=cluster+"/mediaresourcegroup.csv"
cnffile=cluster+"/conferencebridge.csv"
mtpfile=cluster+"/mediaterminationpoint.csv"
txcfile=cluster+"/transcoder.csv"
## Patrones de entrada
mrgname=slc+"-M"

####################### DP
fin = open(dpfile,"r")
csv_f = csv.reader(fin)

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        if row[0].startswith(slc):
            #print(row)
            data['dp'].append({
                'devicepool':row[0].replace('-NOSRST-',''),
                'srst':row[4],
                'mrgl':row[8],
                'region':row[14].replace('-NOSRST-',''),
                'location':row[9].replace('-NOSRST-',''),
                'uso':row[2]
            })
fin.close()

print ("(II): CMO Device Pool",data['dp'], file=f)

####################### LOCATION
fin = open(locationfile,"r")
csv_f = csv.reader(fin)

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        if row[0].startswith(slc):
            #print(row)
            data['loc'].append(({
                'location':row[0].replace('-NOSRST-',''),
                'audio':row[1],
                'video':row[2]
            }))
fin.close()

print ("(II): CMO Location",data['loc'], file=f)

####################### SRST
fin = open(srstfile,"r")
csv_f = csv.reader(fin)

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        if row[0].startswith(slc):
            #print(row)
            data['srst'].append(({
                'ipsccp':row[1],
                'ipsip':row[11]
            }))
fin.close()

print ("(II): CMO SRST",data['srst'], file=f)

####################### Gateway
fin = open(gwfile,"r")
csv_f = csv.reader(fin)
haygw=0

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) > 1: # Me salto las líneas vacias
        if row[4].startswith(slc):
            haygw=1
            data['gw'].append(({
                'trunk':row[0]
            }))

if haygw==0:
    data['gw'].append(({'trunk':"0.0.0.0"}))

fin.close()

print ("(II): CMO GW",data['gw'], file=f)

####################### MRGL
fin = open(mrglfile,"r")
csv_f = csv.reader(fin)

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        if row[0].startswith(mrgname):
            #print(row)
            data['mrgl'].append(({
                'mrg':row[2]
            }))
fin.close()

print("(II) Media Resource Group List: ",data['mrgl'],"(",len(data['mrgl']),")", file=f)
if (len(data['mrgl']) != 1):
    print ("(EE) Es necesario chequear la configuracion...solo puede haber un MRG", file=f)

####################### MRG
fin = open(mrgfile,"r")
csv_f = csv.reader(fin)

## Construyo una lista con lista de recursos
for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        if row[0].startswith(mrgname):
            #print(row)
            for i in row:
                if 'CFB' in i:
                    #print("CFB::",i)
                    data['mrg'].append(i)
                if 'MTP' in i:
                    #print("CFB::",i)
                    data['mrg'].append(i)
                if 'TRANS' in i:
                    #print("CFB::",i)
                    data['mrg'].append(i)

#                if i.startswith(slc) and i.endswith('CFB'): ## Y termina con CFB
#                    print("CFB",i)
#                    data['mrg'].append(i)
#                if i.startswith(slc) and i.find('MTP'): ## Y termina con MTP
#                    print("MTP",i)
#                    data['mrg'].append(i)
#                if i.startswith(slc) and i.find('TRANS'): ## Y termina con TRANS
#                    print("TXC",i)
#                    data['mrg'].append(i)
fin.close()

print("(II) Media Resource Group: ",data['mrg'],"(",len(data['mrg']),")", file=f)
if (len(data['mrg']) != 1):
    print("(EE) Es necesario chequear la configuracion...solo puede haber un MRG ", file=f)

## Si la lista esta vacia salimos no buscamos y seguimos adelante
if len(data['mrg']) == 0:
    print("(EE): No hay recursos MTP, CNF, TXC para SLC=",slc, file=f)
    print("(EE): Algunas pestañas del BLK fallaran!!!!", file=f)
    #exit(1)
else:
    for x in data['mrg']:
        ## Conference
        fin = open(cnffile,"r")
        csv_f = csv.reader(fin)

        for row in csv_f:
            # WR BLK OUTPUT DATA
            if len(row) != 0: # Me salto las líneas vacias
                if row[0] == x:
                    #print(row[0],"=",row[4])
                    data['cnf'].append(({
                        'nombre':row[0],'tipo':row[4]
                    }))
        fin.close()

        ## MTP
        fin = open(mtpfile,"r")
        csv_f = csv.reader(fin)

        for row in csv_f:
            # WR BLK OUTPUT DATA
            if len(row) != 0: # Me salto las líneas vacias
                if row[0] == x:
                    #print(row[0],"=",row[3])
                    data['mtp'].append(({
                        'nombre':row[0],'tipo':row[3]
                    }))
        fin.close()

        # CMO TXC
        fin = open(txcfile,"r")
        csv_f = csv.reader(fin)

        for row in csv_f:
            # WR BLK OUTPUT DATA
            if len(row) != 0: # Me salto las líneas vacias
                if row[0] == x:
                    #print(row[0],"=",row[5])
                    data['trans'].append(({
                        'nombre':row[0],'tipo':row[5]
                    }))
        fin.close()

#################################################################################
#################################################################################
####################### PHONEs
fin = open(phonefile,"r")
#csv_f = csv.reader(fin)
csv_f = csv.DictReader(fin)
intraext1=""
intraext2=""
phonedn=0

#print("(II) ",csv_f.fieldnames,file=f)
header=csv_f.fieldnames

if "Directory Number 1" in header:
    phonedn=1
if "Directory Number 2" in header:
    phonedn=2
if "Directory Number 3" in header:
    phonedn=3
if "Directory Number 4" in header:
    phonedn=4
if "Directory Number 5" in header:
    phonedn=5
if "Directory Number 6" in header:
    phonedn=6

print("(II) El fichero PHONE::DN= ",phonedn,file=f)

phonecount=0

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        ## Nombre Agencia
        if slc in row['Device Pool']:
            phonecount=phonecount+1
            ## INTRASITE-ROUTING
            intraext1,intraext2=busca_digito(row,intraext1,intraext2,phonedn)

print("(II) INTRASITE routing (PHONES): ",intraext1,intraext2, file=f)

fin.close()

#################################################################################
#################################################################################
####################### DEV PROF
fin = open(dproffile,"r")
#csv_f = csv.reader(fin)
csv_f = csv.DictReader(fin)
devprofcount=0
emdn=0

#print("(II) ",csv_f.fieldnames,file=f)
header=csv_f.fieldnames

if "Directory Number 1" in header:
    emdn=1
if "Directory Number 2" in header:
    emdn=2
if "Directory Number 3" in header:
    emdn=3
if "Directory Number 4" in header:
    emdn=4
if "Directory Number 5" in header:
    emdn=5
if "Directory Number 6" in header:
    emdn=6

print("(II) El fichero EM::DN= ",emdn,file=f)

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        ## Nombre Agencia
        if row['Device Profile Name'][1:].startswith(slc):
            devprofcount=devprofcount+1
            ## INTRASITE-ROUTING
            ## INTRASITE-ROUTING
            intraext1,intraext2=busca_digito(row,intraext1,intraext2,emdn)

print("(II) INTRASITE routing (Device Profile): ",intraext1,intraext2, file=f)


if len(intraext1) == 1:
    intraspattern=intraext1
else:
    intraspattern="["+intraext1+"]"

if len(intraext2) == 1:
    intraspattern=intraspattern+intraext2+"XX"
else:
    intraspattern=intraspattern+"["+intraext2+"]XX"


print("(II) IntraSite pattern:",intraspattern, file=f)

## Buscamos una EXTENSION Válida para pruebas
## Por defecto, EXT=9999
siptestext=9999
existe=0
encontrado=1

while (encontrado):
    for row in csv_f:
        # WR BLK OUTPUT DATA
        if len(row) != 0: # Me salto las líneas vacias
            ## Nombre Agencia
            if row['Device Profile Name'][1:].startswith(slc) and row['Directory Number 1'].endswith("9999"):
                existe=1
    if existe == 1: ## Lo hemos buscado y lo hemos encontrado, tenemos que seguir buscando
        encontrado=1
        siptestext=siptestext-1
        print("(II) Buscando "+str(siptestext))
    else:
        print("(II) Buscando "+str(siptestext))
        encontrado=0

print("(II) SIPPTEST extension:",str(siptestext), file=f)


fin.close()


####################### E164 -- Mask
# CMO File INPUT DATA
#cmositedata = "../CMO/AGENCIAS CISCO - MIGRAÇÃO HCS.xlsx"
cmositedata = "../CMO/AGENCIAS-CISCO_HCS.xlsx"
blk = openpyxl.load_workbook(cmositedata,read_only=True)
sheet = blk["Plan1"]

for row in sheet.rows:
    #print("<<>>",row[0].value)
    #DEBUG
    #print(sheet.cell(row=fila,column=col).value)
    rslc=str(row[0].value)
    if rslc[1:].startswith(slc):
        ## AREACODE + CABECERA
        sitename=slc+"-"+row[1].value
        areacode=row[13].value[1:3]
        nacional=row[13].value[5:]
        ramais=row[5].value
        uniorg=row[0].value
        ##
        epnm="+55"+areacode+nacional
        print("(II) SLC encontrado: \t\t",rslc, file=f)
        print("(II) SITE NAME: \t\t",sitename, file=f)
        print("(II) CABECERA: \t\t",row[13].value, file=f)
        print("(II) AREA CODE: \t\t",areacode, file=f)
        print("(II) CABECERA +E164: \t",epnm, file=f)
        print("(II) RAMAIS : \t\t\t",ramais, file=f)
        print("(II) PHONE COUNT: \t\t",phonecount, file=f)
        print("(II) Dev Prof COUNT: \t\t",devprofcount, file=f)
#    else:
#        ## CONDICION de ERROR
#        ##
#        print("(EE): NO EXISTE el SITE ",slc," en el fichero de datos",cmositedata)

if sitename == "" or areacode == "" or epnm == "" or ramais == "":
    print("(EE): NO EXISTE el SITE ",slc," en el fichero de datos",cmositedata, file=f)
    print("(EE): Es necesario actualizar el fichero ",cmositedata, file=f)
    print("(EE): Y volver a ejecutar el comando ", file=f)

#blk.save(cmositedata)  ## blk = openpyxl.load_workbook(cmositedata,read_only=True) -> Read_only switch

## DEBUG
#print(">>>>> ",phonecount,devprofcount)
data['devices'].append({'phones':phonecount,'udp':devprofcount,'ramais':ramais})
data['e164'].append({'epnm':"",'head':epnm,'slc':slc,'ac':areacode,'cc':"+55",'intrasite1':intraext1,'intrasite2':intraext2,'patternintra':intraspattern,'phonedn':phonedn,'emdn':emdn})
## FMOSITE
data['fmosite'].append({'name':sitename,'id':"",'cmg':"",'uniorg':uniorg})
data['sipptest'].append({'extension':str(siptestext)})


with open(outconfigfile,'w') as outfile:
    json.dump(data,outfile)
outfile.close()

## LOG de CONFIGURACION
f.close()

exit(0)
