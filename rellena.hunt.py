## LINUX#!/usr/bin/python3
## OSX
#!/usr/local/bin/python3

import csv
import openpyxl
import time
import sys
import json

#INPUTS: SLC
siteslc="0000"

if len(sys.argv) != 6:
    print("ERROR: python3 <nombre-fichero.py> <static-data-file> <site-data-file> XXXX <ClusterPath> <Config Log File>")
    exit(1)
else:
    fmostaticdata=sys.argv[1]
    fmositedata=sys.argv[2]
    siteslc=str(sys.argv[3])   #INPUTS: SLC
    clusterpath=sys.argv[4]    #SLC Cluster path
    logconfigfile=sys.argv[5]   ## LOG config file

## LOG File
f = open(logconfigfile, 'a')

print("(II) #################################################", file=f)
print("(II) #################################################", file=f)
print("(II) INPUT CMO configuration      ::",fmositedata, file=f)
print("(II) INPUT datos de entorno       ::",fmostaticdata, file=f)
print("(II) INPUT SLC                    :: ",siteslc, file=f)
print("(II) INPUT LOG configuration file :: ",logconfigfile, file=f)
print("(II) INPUT Cluster Path           :: ",clusterpath, file=f)

cl=clusterpath[14:16]       ## CL = dos digitos, 01, 02, 03,...

# FMO datos de entorno:
fmoenvconfig={}
with open(fmostaticdata, "r") as fp:
        for line in fp.readlines():
            li = line.lstrip()
            if not li.startswith("#") and '=' in li:
                key, value = line.split('=', 1)
                fmoenvconfig[key] = value.strip()  ## variable de tipo diccionario
                #print("<<<<   ",fmoenvconfig[key])
fp.close()

# CMO datos del site
data = {}
#data['dp'] = []
#data['loc']=[]
#data['mrgl'] = []
#data['mrg'] = []
#data['srst'] = []
#data['gw'] = []
#data['rsc'] = []
#data['e164'] = []
#data['agencia'] = []


#e164={}
#agencia={}

with open(fmositedata,'r') as infile:
    data = json.load(infile)
infile.close()

#FMO working path:
sitepath="../FMO/"+siteslc

# FILES
inputfilehl = clusterpath+"/huntlist.csv"       ## ORIGINAL
inputfilehp = clusterpath+"/huntpilot.csv"      ## ORIGINAL
#inputfilelg = clusterpath+"/linegroup.csv"      ## ORIGINAL ->
inputfilelg = clusterpath+"/linegroup.mod2.csv"      ## MOD

templateblkfile = "blk/06.hunt-template.xlsx" # SIN DATAINPUT
outputblkfile = sitepath+"/06.hunt."+siteslc+".xlsx"

## FMO CUSTOMER INPUT DATA
hierarchynode=fmoenvconfig['hierarchynode']
customerid=fmoenvconfig['fmocustomerid']
aargroup=fmoenvconfig['fmoaargroup']
##
fmositename=data['fmosite']['name']
fmositeid=data['fmosite']['id']
cmg=data['fmosite']['cmg']

# CMO patterns
cmodevicepool=siteslc+"-DP"

## FMO UserData
cucdmsite=fmoenvconfig['fmocustomerid']+"Si"+str(fmositeid)
cssfwd=customerid+"-DirNum-CSS"
linept=customerid+"-DirNum-PT"
emlinept=customerid+"-DirNumEM-PT"
linecss=cucdmsite+"-DBRSTDNatl24HrsCLIPyFONnFACnCMC-CSS"
aarcss=customerid+"-AAR-CSS"
devicepool=cucdmsite+"-DevicePool"
location=cucdmsite+"-Location"
devicecss=cucdmsite+"-BRADP-DBRDevice-CSS"
subscribecss=cucdmsite+"-InternalOnly-CSS"
preisrpt=customerid+"-PreISR-PT"
preisrcss=customerid+"-PreISR-CSS"
isrpt=customerid+"-ISR-PT"
isrcss=customerid+"-ISR-CSS"
dirnumpt=customerid+"-DirNum-PT"
dirnumcss=customerid+"-DirNum-CSS"
fwdhuntcss=cucdmsite+"-HPFWD-CSS"

# CMO File INPUT DATA
fgw = open(inputfilehp,"r")
csv_f = csv.DictReader(fgw)
#csv_f = csv.reader(fgw)

# FMO File OUTPUT DATA
blk = openpyxl.load_workbook(templateblkfile)

# FMO commands:
action=""

fila=7
hunt=[]
huntpilot=[]
linegroup=[]

######################################################
## CMO HUNT PILOT
######################################################
print("(II): CMO Hunt Pilots site",siteslc,file=f)

for row in csv_f:
    # WR BLK OUTPUT DATA
    if len(row) != 0: # Me salto las líneas vacias
        if row['HUNT PILOT'][1:].startswith(siteslc):
            ## DEBUG
            #print("HL#",fila,row['HUNT PILOT']," ##L",fila,"##: ",file=f)
            ######################################################
            ## HUNT PILOT
            ######################################################
            sheet =  blk["HP"]
            sheet['B'+str(fila)]=hierarchynode+"."+fmositename
            sheet['C'+str(fila)]=action
            #sheet['D'+str(fila)]="name:"+row[3] # Search field
            ######################################################
            sheet['i'+str(fila)]=row['PARK MONITOR FORWARD NO RETRIEVE DESTINATION']                               # parkMonForwardNoRetrieve.destination
            sheet['j'+str(fila)]="false"                               # parkMonForwardNoRetrieve.usePersonalPreferences
            sheet['k'+str(fila)]=fwdhuntcss                                    # #parkMonForwardNoRetrieve.callingSearchSpaceName
            sheet['l'+str(fila)]=row['ASCII ALERTING NAME']                                # asciiAlertingName
            sheet['m'+str(fila)]=""                                    # callingPartyPrefixDigits
            sheet['n'+str(fila)]="Cisco CallManager"                   # callingPartyNumberingPlan
            sheet['o'+str(fila)]="Default"                             # connectedLinePresentationBit
            #sheet['p'+str(fila)]=row[]                                 # CPG
            sheet['q'+str(fila)]=linept                                 # routePartitionName
            sheet['r'+str(fila)]="No Error"                             # releaseClause
            sheet['s'+str(fila)]="false"                                # displayConnectedNumber
            sheet['t'+str(fila)]=row['DESCRIPTION']                                 # description
            sheet['u'+str(fila)]="Cisco CallManager"                    # callingPartyNumberType
            sheet['v'+str(fila)]="true"                                 # provideOutsideDialtone
            sheet['w'+str(fila)]=""                                     # callingPartyTransformationMask
            sheet['x'+str(fila)]=row['HUNT PILOT']                      # pattern
            sheet['y'+str(fila)]=""                                     #patternPrecedence
            sheet['z'+str(fila)]=""                                    # prefixDigitsOut
            sheet['aa'+str(fila)]=""                                     # #maxHuntduration
            sheet['ab'+str(fila)]="Cisco CallManager"                   # calledPartyNumberingPlan
            ##
            sheet['ac'+str(fila)]=row['HUNT LIST 1']                               # HuntList.name
            ##
            sheet['ad'+str(fila)]=row['ALERTING NAME']                                # alertingName
            sheet['ae'+str(fila)]="Default"                             # connectedLinePresentationBit
            sheet['s'+str(fila)]="false"                                # blockEnable
            sheet['ag'+str(fila)]=""                                    # callingNamePresentationBit
            sheet['ah'+str(fila)]=""                                    #dialplanName
            sheet['ai'+str(fila)]=""                                    # e164Mask
            sheet['aj'+str(fila)]=""                                    # digitDiscardInstructionName
            sheet['ak'+str(fila)]="Off"                                 # useCallingPartyPhoneMask
            sheet['al'+str(fila)]=""                                    # callingPartyTransformationMask
            sheet['am'+str(fila)]="Cisco CallManager"                   # calledPartyNumberType
            sheet['an'+str(fila)]=""                                    #routefilterName
            if row['FORWARD HUNT NO ANSWER DESTINATION'] == "NULL" or row['FORWARD HUNT NO ANSWER DESTINATION']=="":
                sheet['ao'+str(fila)]=""
            else:
                sheet['ao'+str(fila)]=row['FORWARD HUNT NO ANSWER DESTINATION']                                # forwardHuntNoAnswer.destination
            sheet['ap'+str(fila)]=""                                     # forwardHuntNoAnswer.usePersonalPreferences
            sheet['aq'+str(fila)]=fwdhuntcss                                   # #forwardHuntNoAnswer.callingSearchSpaceName
            sheet['ar'+str(fila)]=""                                     # patternUrgency
            sheet['as'+str(fila)]=fmoenvconfig['fmoaargroup']           # aarNeighborhoodName
            if row['FORWARD HUNT BUSY DESTINATION'] == "NULL" or row['FORWARD HUNT BUSY DESTINATION']=="":
                sheet['at'+str(fila)]=""
            else:
                sheet['at'+str(fila)]=row['FORWARD HUNT BUSY DESTINATION']                               # forwardHuntBusy.destination
            sheet['au'+str(fila)]="false"                               # forwardHuntBusy.usePersonalPreferences
            sheet['av'+str(fila)]=fwdhuntcss                                  # #forwardHuntBusy.callingSearchSpaceName
            sheet['aw'+str(fila)]="Default"                             # callingLinePresentationBit

            ## Tabla con la cadena completa
            hunt.append(row['HUNT LIST 1'])
            huntpilot.append(row['HUNT PILOT']) ## Lo guardamos para desactivar/activar los patrones

            ######################################################
            ######################################################
            ######################################################
            ## Creamos el routing para los HP PreISR
            sheet = blk["TPpreISR"]
            sheet['B'+str(fila)]=hierarchynode+"."+fmositename
            sheet['C'+str(fila)]=action
            #sheet['D'+str(fila)]="name:"+row[3]
            sheet['I'+str(fila)]="Cisco CallManager"           # callingPartyNumberingPlan
            sheet['j'+str(fila)]="Default"                     # connectedLinePresentationBit
            sheet['k'+str(fila)]=preisrpt                      # routePartitionName
            sheet['l'+str(fila)]="No Error"                    # releaseClause
            sheet['m'+str(fila)]="false"                       # blockEnable
            sheet['n'+str(fila)]="Cisco CallManager"           # callingPartyNumberType
            sheet['o'+str(fila)]="false"                       # provideOutsideDialtone
            sheet['q'+str(fila)]=row['HUNT PILOT']             # pattern
            sheet['r'+str(fila)]="Default"                     # patternPrecedence
            sheet['s'+str(fila)]=isrcss                        # callingSearchSpaceName
            sheet['t'+str(fila)]=""                            # prefixDigitsOut
            sheet['u'+str(fila)]="Translation"                 # usage
            sheet['v'+str(fila)]="Cisco CallManager"           # calledPartyNumberingPlan
            sheet['w'+str(fila)]="true"                        # dontWaitForIDTOnSubsequentHops
            sheet['x'+str(fila)]="Default"                     # connectedNamePresentationBit
            sheet['y'+str(fila)]="PreISR HP "+row['HUNT LIST 1']+row['HUNT PILOT']   # Description
            sheet['z'+str(fila)]="Default"                     # routeClass
            sheet['aa'+str(fila)]="Default"                    # callingNamePresentationBit
            sheet['ac'+str(fila)]="false"                      # routeNextHopByCgpn
            sheet['ad'+str(fila)]="false"                       # useOriginatorCss
            sheet['ae'+str(fila)]=""                           # digitDiscardInstructionName
            sheet['ag'+str(fila)]="Off"                        # useCallingPartyPhoneMask
            sheet['aj'+str(fila)]="Cisco CallManager"          # calledPartyNumberType
            sheet['al'+str(fila)]="false"                      # patternUrgency
            sheet['am'+str(fila)]="Default"                    # callingLinePresentationBit

            ######################################################
            ######################################################
            ######################################################
            ## Creamos el routing para los HP PreISR
            sheet = blk["TPISR"]
            sheet['B'+str(fila)]=hierarchynode+"."+fmositename
            sheet['C'+str(fila)]=action
            #sheet['D'+str(fila)]="name:"+row[3]
            sheet['I'+str(fila)]="Cisco CallManager"           # callingPartyNumberingPlan
            sheet['j'+str(fila)]="Default"                     # connectedLinePresentationBit
            sheet['k'+str(fila)]=isrpt                      # routePartitionName
            sheet['l'+str(fila)]="No Error"                    # releaseClause
            sheet['m'+str(fila)]="false"                       # blockEnable
            sheet['n'+str(fila)]="Cisco CallManager"           # callingPartyNumberType
            sheet['o'+str(fila)]="false"                       # provideOutsideDialtone
            sheet['q'+str(fila)]=row['HUNT PILOT']                        # pattern
            sheet['r'+str(fila)]="Default"                     # patternPrecedence
            sheet['s'+str(fila)]=dirnumcss                        # callingSearchSpaceName
            sheet['t'+str(fila)]=""                            # prefixDigitsOut
            sheet['u'+str(fila)]="Translation"                 # usage
            sheet['v'+str(fila)]="Cisco CallManager"           # calledPartyNumberingPlan
            sheet['w'+str(fila)]="true"                        # dontWaitForIDTOnSubsequentHops
            sheet['x'+str(fila)]="Default"                     # connectedNamePresentationBit
            sheet['y'+str(fila)]="ISR CPG "+row['HUNT LIST 1']+row['HUNT PILOT']      # Description
            sheet['z'+str(fila)]="Default"                     # routeClass
            sheet['aa'+str(fila)]="Default"                    # callingNamePresentationBit
            sheet['ac'+str(fila)]="false"                      # routeNextHopByCgpn
            sheet['ad'+str(fila)]="false"                       # useOriginatorCss
            sheet['ae'+str(fila)]=""                           # digitDiscardInstructionName
            sheet['ag'+str(fila)]="Off"                        # useCallingPartyPhoneMask
            sheet['aj'+str(fila)]="Cisco CallManager"          # calledPartyNumberType
            sheet['al'+str(fila)]="false"                      # patternUrgency
            sheet['am'+str(fila)]="Default"                    # callingLinePresentationBit

            fila+=1

## CMO File INPUT DATA: Close
fgw.close()

## DEBUG
print("(II) HUNT LIST NAMES:",hunt,file=f)
print("(II) HUNT LIST NUMBERS:",huntpilot,file=f)
## Guardo los Hunt pilot patterns
with open(fmositedata,'w') as outfile:
    #data = json.load(outfile)
    data['huntpilot']=huntpilot
    json.dump(data,outfile)
outfile.close()

######################################################
## CMO HUNT LIST
######################################################
print("(II): CMO Hunt Lists site",siteslc,file=f)

# CMO File INPUT DATA
fgw = open(inputfilehl,"r")
#csv_f = csv.reader(fgw)
csv_f = csv.DictReader(fgw)

fila=7

for row in csv_f:           ## WR BLK OUTPUT DATA
    for hlname in hunt:
        if len(row) != 0:   ## Me salto las líneas vacias
            if row['NAME'] == hlname:
                ######################################################
                ## HUNT LIST
                ######################################################
                sheet =  blk["HL"]
                sheet['B'+str(fila)]=hierarchynode+"."+fmositename
                sheet['C'+str(fila)]=action
                #sheet['D'+str(fila)]="name:"+row[3] # Search field
                ######################################################

                sheet['i'+str(fila)]="false"                                # voiceMailUsage
                sheet['j'+str(fila)]=cmg                                    # callManagerGroupName
                sheet['k'+str(fila)]=row['NAME']                            # name @@@@@@
                sheet['l'+str(fila)]=row['SELECTION ORDER 1']               # members.member.0.selectionOrder
                sheet['m'+str(fila)]=row['LINE GROUP 1']                    # members.member.0.lineGroupName
                sheet['n'+str(fila)]="true"                                 # routeListEnabled
                linegroup.append(row['LINE GROUP 1'])
                ##
                if len(row) > 11 and row['SELECTION ORDER 2'] != "":
                    sheet['o'+str(fila)]=row['LINE GROUP 2']                            # members.member.1.selectionOrder
                    sheet['p'+str(fila)]=row['SELECTION ORDER 2']                            # members.member.1.lineGroupName
                    linegroup.append(row['LINE GROUP 2'])
                ##
                if len(row) > 13 and row['SELECTION ORDER 3'] != "":
                    sheet['q'+str(fila)]=row['SELECTION ORDER 3']                            # members.member.2.selectionOrder
                    sheet['r'+str(fila)]=row['LINE GROUP 3']                            # members.member.2.lineGroupName
                    linegroup.append(row['LINE GROUP 3'])
                ##
                if len(row) > 15 and row['SELECTION ORDER 4'] != "":
                    sheet['s'+str(fila)]=row['SELECTION ORDER 5']                            # members.member.3.selectionOrder
                    sheet['t'+str(fila)]=row['LINE GROUP 4']                            # members.member.3.lineGroupName
                    linegroup.append(row['LINE GROUP 4'])
                ##
                if len(row) > 17 and row['SELECTION ORDER 5'] != "":
                    sheet['u'+str(fila)]=row['SELECTION ORDER 5']                            # members.member.4.selectionOrder
                    sheet['v'+str(fila)]=row['LINE GROUP 5']                            # members.member.4.lineGroupName
                    linegroup.append(row['LINE GROUP 5'])

                fila+=1

## CMO File INPUT DATA: Close
fgw.close()

## DEBUG
print("(II) LINE GROUP ::",linegroup,file=f)

######################################################
## CMO LINE GROUP
######################################################
print("(II) CMO Line Group site",siteslc,file=f)

# CMO File INPUT DATA
fgw = open(inputfilelg,"r")
csv_f = csv.reader(fgw)
#csv_f = csv.DictReader(fgw) ## NO SE PUEDE USAR

# FMO LG header
fila=2
col=16 ## Empezamos en la columna 16 ## "P"
lgpos=0
sheet =  blk["LG"]
while lgpos < 81:   ## Max num patterns
    ## Terna de valores que se repite
    #sheet['m'+str(fila)]=row[]                   ## M=12        # members.member.0.directoryNumber.pattern
    #sheet['n'+str(fila)]=""                      ## N=13        # members.member.0.directoryNumber.routePartitionName
    #sheet['o'+str(fila)]=row                     ## N=14        # members.member.0.lineSelectionOrder
    #print("[",fila,"][",col,"]",file=f)
    sheet.cell(row=fila,column=col).value="members.member."+str(lgpos)+".lineSelectionOrder"
    col+=1
    sheet.cell(row=fila,column=col).value="members.member."+str(lgpos)+".directoryNumber.pattern"
    col+=1
    sheet.cell(row=fila,column=col).value="members.member."+str(lgpos)+".directoryNumber.routePartitionName"
    col+=1
    lgpos+=1

# FMO customer data
fila=7

for row in csv_f:               ## WR BLK OUTPUT DATA
    for lgname in linegroup:    ##
        if len(row) != 0:       ## Me salto las líneas vacias
            if row[1] == lgname:
                #print("Buscando ",lgname," en",row[1],file=f)
                #####################################################
                ## LINE GROUP
                ######################################################
                sheet =  blk["LG"]
                sheet['B'+str(fila)]=hierarchynode+"."+fmositename
                sheet['C'+str(fila)]=action
                #sheet['D'+str(fila)]="name:"+row[3] # Search field
                ######################################################

                sheet['j'+str(fila)]=row[1]                             # name
                sheet['k'+str(fila)]=row[2]      # distributionAlgorithm
                sheet['m'+str(fila)]=row[0]            # rnaReversionTimeOut
                sheet['o'+str(fila)]="false"                                 # autoLogOffHunt

                if "Hunt" in row[3]:                 # huntAlgorithmNoAnswer -> CMO #3
                    sheet['i'+str(fila)]="Try next member; then, try next group in Hunt List"
                else:
                    sheet['i'+str(fila)]="Try next member, but do not go to next group"

                if "Hunt" in row[5]:                # huntAlgorithmNotAvailable -> CMO #5
                    sheet['l'+str(fila)]="Try next member; then, try next group in Hunt List"
                else:
                    sheet['l'+str(fila)]="Try next member, but do not go to next group"

                if "Hunt" in row[4]:                # huntAlgorithmBusy -> CMO #4
                    sheet['n'+str(fila)]="Try next member; then, try next group in Hunt List"
                else:
                    sheet['n'+str(fila)]="Try next member, but do not go to next group"

                ## Terna de valores que se repite
                #sheet[''+str(fila)]=row[]                   ## M=12        # members.member.0.directoryNumber.pattern
                #sheet[''+str(fila)]=""                      ## N=13        # members.member.0.directoryNumber.routePartitionName
                #sheet[''+str(fila)]=row                     ## N=14        # members.member.0.lineSelectionOrder

                colnum=1
                colvalid=16
                for col in row:
                    if colnum > 6 and col != "":
                    #if colnum > 6:
                        #print("[",fila,"][",colvalid,"]=",col,file=f)
                        if col == "Interna-PT":
                            sheet.cell(row=fila,column=colvalid).value=emlinept
                        elif col == "Interna-EM-PT":
                            sheet.cell(row=fila,column=colvalid).value=linept
                        else:
                            sheet.cell(row=fila,column=colvalid).value=col
                        colvalid+=1
                    colnum+=1

                fila+=1

## CMO File INPUT DATA: Close
fgw.close()

## FMO File OUTPUT DATA: Close
blk.save(outputblkfile)

## LOG de CONFIGURACION
f.close()

exit(0)
